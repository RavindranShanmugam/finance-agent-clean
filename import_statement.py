"""
import_statement.py  -  Build/update a personal spending tracker from Bank of America PDF statements.

HOW TO USE
  1. Drop a statement PDF in this folder.
  2. Run:   python import_statement.py  yourstatement.pdf
            (or just  python import_statement.py  to grab the newest PDF in the folder)
  3. It adds the new transactions into  spending.xlsx  (creating it the first time),
     skips any it already imported, and recomputes your Top Spending.
  Run it again each month with a new statement -> the sheet grows, the totals update.
"""
import sys, re, glob, os
import pdfplumber
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

MASTER = "spending.xlsx"

# ---- CATEGORY RULES: add your own merchants here over time ----
CATEGORY_RULES = {
    "Coffee & Cafe":      ["FOUNT COFFEE", "CARIBOU", "STARBUCKS", "PANERA"],
    "Fast Food / Dining": ["SHAWARMA", "SUBWAY", "MCDONALD", "CHIPOTLE", "SANGAM", "BROGDEN"],
    "Groceries":          ["TARGET", "WAL-MART", "WALMART", "FOOD LION", "LOWES FOODS"],
    "Gas & Convenience":  ["CIRCLE K", "BP#", "GIENC", "NAYAX"],
    "Shopping / Retail":  ["TJMAXX", "TJ MAXX"],
    "Software & Subs":    ["HOSTINGER", "OPENROUTER", "OPENAI", "CHATGPT", "TWILIO"],
    "Shipping":           ["UPS STORE"],
    "Phone":              ["CONS PHN PMT", "PHN PMT"],
    "Investing":          ["ROBINHOOD", "ACORNS"],
    "Transfers / Other":  ["ZELLE", "CAPITAL ONE", "BKOFAMERICA BC", "RMTLY", "WITHDRWL", "PMNT SENT"],
    "Fees":               ["OVERDRAFT", "NSF"],
}
NON_SPEND = ["Investing", "Transfers / Other", "Fees"]   # money out, but not "spending"

def categorize(desc):
    up = desc.upper()
    for cat, kws in CATEGORY_RULES.items():
        if any(k in up for k in kws):
            return cat
    return "Uncategorized"

def parse_statement(pdf_path):
    rows, section = [], None
    date_re = re.compile(r"^(\d{2}/\d{2}/\d{2})\s+(.*)")
    amt_re  = re.compile(r"(-?[\d,]+\.\d{2})\s*$")
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            for line in (page.extract_text() or "").split("\n"):
                s = line.strip()
                if "Deposits and other additions" in s: section = "Income"; continue
                if "ATM and debit card subtractions" in s: section = "Spending"; continue
                if "Other subtractions" in s: section = "Other"; continue
                m = date_re.match(s)
                if not m: continue
                a = amt_re.search(s)
                if not a: continue
                rows.append({
                    "Date": m.group(1),
                    "Description": s[len(m.group(1)):a.start()].strip(),
                    "Amount": float(a.group(1).replace(",", "")),
                    "Source": os.path.basename(pdf_path),
                })
    return pd.DataFrame(rows)

def autoformat(path):
    """Make the Excel look clean: bold headers, currency, sized columns."""
    import openpyxl
    wb = openpyxl.load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        # size columns + currency format
        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[letter].width = min(width + 3, 60)
            header = ws.cell(row=1, column=col[0].column).value
            if header in ("Amount", "Spent"):
                for c in col[1:]:
                    c.number_format = '$#,##0.00'
        ws.freeze_panes = "A2"
    wb.save(path)

def main():
    pdf = sys.argv[1] if len(sys.argv) > 1 else max(glob.glob("*.pdf"), key=os.path.getmtime)
    new = parse_statement(pdf)
    new["Category"] = new["Description"].apply(categorize)

    # merge with whatever is already in the master file
    if os.path.exists(MASTER):
        old = pd.read_excel(MASTER, sheet_name="Transactions")
        combined = pd.concat([old, new], ignore_index=True)
    else:
        combined = new
    # de-duplicate so re-importing the same statement does nothing
    combined = combined.drop_duplicates(subset=["Date", "Description", "Amount"]).reset_index(drop=True)

    # build the Top Spending summary (real spending only)
    spend = combined[(combined["Amount"] < 0) & (~combined["Category"].isin(NON_SPEND))].copy()
    spend["Spent"] = spend["Amount"].abs()
    summary = spend.groupby("Category")["Spent"].sum().sort_values(ascending=False).reset_index()
    total = pd.DataFrame([{"Category": "TOTAL", "Spent": summary["Spent"].sum()}])
    summary = pd.concat([summary, total], ignore_index=True)

    with pd.ExcelWriter(MASTER, engine="openpyxl") as xl:
        combined.to_excel(xl, sheet_name="Transactions", index=False)
        summary.to_excel(xl, sheet_name="Top Spending", index=False)
    autoformat(MASTER)

    print(f"Imported {len(new)} transactions from {os.path.basename(pdf)}")
    print(f"{MASTER} now holds {len(combined)} transactions total.\n")
    print(summary.to_string(index=False))

if __name__ == "__main__":
    main()
